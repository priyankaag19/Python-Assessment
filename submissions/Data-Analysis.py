import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

def reverse_by_n_elements(lst, n):
    """Reverse the list in chunks of n elements."""
    return [lst[i:i + n][::-1] for i in range(0, len(lst), n)]

def group_by_length(strings):
    """Group strings by their length."""
    grouped = {}
    for string in strings:
        length = len(string)
        if length not in grouped:
            grouped[length] = []
        grouped[length].append(string)
    return grouped

def flatten_dict(nested_dict, parent_key='', sep='_'):
    """Flatten a nested dictionary."""
    items = []
    for k, v in nested_dict.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            # Convert lists to strings
            items.append((new_key, str(v)))  # Convert list to string for Excel
        else:
            items.append((new_key, v))
    return dict(items)

def create_excel_file():
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Analysis"

    # Example data from the functions
    example_list = [1, 2, 3, 4, 5, 6, 7, 8]
    reversed_list = reverse_by_n_elements(example_list, 3)
    
    example_strings = ["apple", "bat", "car", "elephant", "dog", "bear"]
    grouped_strings = group_by_length(example_strings)

    # Flattened dictionary example
    nested_dict = {
        "road": {
            "name": "Highway 1",
            "length": 350,
            "sections": [
                {
                    "id": 1,
                    "condition": {
                        "pavement": "good",
                        "traffic": "moderate"
                    }
                }
            ]
        }
    }
    flattened_dict = flatten_dict(nested_dict)

    # Write data to Excel
    ws.append(["Reversed List"])
    for chunk in reversed_list:
        ws.append(chunk)

    # Write grouped strings to Excel
    ws.append([])
    ws.append(["Grouped Strings"])
    for length, strings in grouped_strings.items():
        ws.append([length, ', '.join(strings)])

    # Write flattened dictionary to Excel
    ws.append([])
    ws.append(["Flattened Dictionary"])
    for key, value in flattened_dict.items():
        ws.append([key, value])

    # Apply conditional formatting: Fill color for the reversed list
    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    for row in ws.iter_rows(min_row=2, max_row=len(reversed_list)+1, min_col=1, max_col=len(reversed_list[0])):
        for cell in row:
            cell.fill = fill

    # Set column widths
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Save the workbook
    wb.save("excel-assessment.xlsm")

# Example function definitions to run
if __name__ == "__main__":
    create_excel_file()
    print("Excel file created successfully.")
